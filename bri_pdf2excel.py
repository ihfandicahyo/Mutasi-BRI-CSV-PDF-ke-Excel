import pdfplumber
import pandas as pd
import re
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- KONFIGURASI ---
# Pola Regex untuk mendeteksi tanggal di awal baris (dd/mm/yy)
DATE_PATTERN = re.compile(r'^(\d{2}/\d{2}/\d{2})')

def format_excel_output(filename):
    """
    Melakukan Autofit lebar kolom 
    Mengubah format kolom E, F, G menjadi format ribuan (Accounting/Number).
    """
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        # 1. AUTOFIT (Menyesuaikan lebar kolom)
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length: max_length = length
                except: pass
            # Tambah sedikit padding agar rapi
            ws.column_dimensions[column_letter].width = max_length + 2

        # 2. FORMAT RIBUAN INDONESIA (Kolom E, F, G)
        # E=Debet, F=Kredit, G=Saldo
        # Di OpenPyXL, kolom diakses dengan huruf 'E', 'F', 'G'
        target_columns = ['E', 'F', 'G']
        
        for col_letter in target_columns:
            # Loop setiap sel di kolom tersebut (mulai baris 2 untuk skip header)
            for cell in ws[col_letter]:
                if cell.row > 1: # Skip Header
                    # Set format angka dengan pemisah ribuan
                    # '#,##0' akan otomatis menjadi 1.000.000 di Excel Region Indonesia
                    # atau 1,000,000 di Excel Region US.
                    cell.number_format = '#,##0' 

        wb.save(filename)
    except Exception as e:
        print(f"   [!] Warning saat formatting Excel: {e}")

def parse_line_brimo(line):
    """Memecah baris teks menjadi data transaksi."""
    parts = line.split()
    
    # 1. Ambil Tanggal
    txn_date = parts[0]
    
    # 2. Ambil Jam (jika ada)
    current_idx = 1
    txn_time = ""
    if len(parts) > 1 and ':' in parts[1]:
        txn_time = parts[1]
        current_idx = 2
        
    # 3. Ambil Angka (Debet, Kredit, Saldo) dari BELAKANG (Reverse Parsing)
    saldo = "0"
    kredit = "0"
    debet = "0"
    teller_id = ""
    
    try:
        saldo = parts[-1]   # Paling kanan
        kredit = parts[-2]  # Sebelah kiri saldo
        debet = parts[-3]   # Sebelah kiri kredit
        
        # Cek Teller ID (biasanya angka di sebelah kiri Debet)
        remainder_end_idx = -3
        if len(parts) > 4 and parts[-4].replace('.', '').isdigit():
             teller_id = parts[-4]
             remainder_end_idx = -4
             
        # Sisanya di tengah adalah Uraian
        uraian_parts = parts[current_idx:remainder_end_idx]
        uraian = " ".join(uraian_parts)
        
    except IndexError:
        return None

    return {
        "Tanggal": txn_date,
        "Jam": txn_time,
        "Uraian": uraian,
        "Teller": teller_id,
        "Debet": debet,
        "Kredit": kredit,
        "Saldo": saldo
    }

def process_pdf(pdf_path, output_path):
    filename = os.path.basename(pdf_path)
    print(f"\n-> Memproses: {filename}")
    
    transactions = []
    inside_table = False
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                # Ekstrak text layout-preserving
                text = page.extract_text(x_tolerance=2, y_tolerance=3)
                if not text: continue
                
                lines = text.split('\n')
                for line in lines:
                    line = line.strip()
                    if not line: continue
                    
                    # Deteksi Awal & Akhir Tabel
                    if "Tanggal Transaksi" in line or "Transaction Date" in line:
                        inside_table = True
                        continue
                    if "Saldo Awal" in line or "Opening Balance" in line or "Total Transaksi" in line:
                        inside_table = False
                        continue
                        
                    if inside_table:
                        # Jika baris dimulai dengan Tanggal -> Transaksi Baru
                        if DATE_PATTERN.match(line):
                            parsed = parse_line_brimo(line)
                            if parsed:
                                transactions.append(parsed)
                        else:
                            # Jika tidak ada tanggal -> Lanjutan Uraian baris sebelumnya
                            if transactions:
                                transactions[-1]["Uraian"] += " " + line

        if not transactions:
            print(f"   [!] Data kosong/tidak terbaca: {filename}")
            return

        # Buat DataFrame
        df = pd.DataFrame(transactions)
        
        # Pastikan urutan kolom sesuai untuk Excel (A,B,C,D,E,F,G)
        cols_order = ["Tanggal", "Jam", "Uraian", "Teller", "Debet", "Kredit", "Saldo"]
        # Filter hanya kolom yang berhasil di-parse
        final_cols = [c for c in cols_order if c in df.columns]
        df = df[final_cols]

        # Cleaning Data Angka (Hapus 'IDR', 'Rp', koma ribuan)
        for col in ["Debet", "Kredit", "Saldo"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.replace(',', '', regex=False)
                df[col] = df[col].str.replace('IDR', '', regex=False)\
                                 .str.replace('Rp', '', regex=False)\
                                 .str.strip()
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Simpan ke Excel
        df.to_excel(output_path, index=False)
        
        # Terapkan Formatting (Autofit & Format Ribuan)
        format_excel_output(output_path)
        
        print(f"   [V] Sukses! Output: {os.path.basename(output_path)}")

    except Exception as e:
        print(f"   [X] Error: {e}")

def main():
    # Deteksi folder tempat script berada
    if getattr(sys, 'frozen', False):
        app_path = os.path.dirname(sys.executable)
    else:
        app_path = os.path.dirname(os.path.abspath(__file__))

    print("="*60)
    print(f"AUTO CONVERTER PDF BANK BRI (Scan Folder)")
    print(f"Lokasi: {app_path}")
    print("="*60)
    
    # Cari semua file .pdf di folder tersebut
    files = [f for f in os.listdir(app_path) if f.lower().endswith('.pdf')]
    
    if not files:
        print("TIDAK ADA FILE PDF DI FOLDER INI.")
    else:
        print(f"Ditemukan {len(files)} file PDF. Memulai proses...\n")
        count = 0
        for f in files:
            pdf_path = os.path.join(app_path, f)
            # Nama output sama dengan input tapi .xlsx
            xlsx_path = os.path.join(app_path, os.path.splitext(f)[0] + ".xlsx")
            
            process_pdf(pdf_path, xlsx_path)
            count += 1
            
    print("\n" + "="*60)
    print(f"Proses Selesai. {count} file telah diproses.")
    input("Tekan Enter untuk menutup...")

if __name__ == "__main__":
    main()